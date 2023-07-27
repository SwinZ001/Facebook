import openpyxl as openpyxl
import requests
# import sys
# reload(sys)
# sys.setdefaultencoding( "utf-8" )

if __name__ == "__main__":
    url = 'https://erp-product.heidongwl.com/api/cabcli/public_auth_data'
    headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.104 Safari/537.36'
    }
    #
    xl=openpyxl.load_workbook(r'G:\test.xlsx')
    sheets_name=xl["Sheet"]
    x = 1
    #
    wb = openpyxl.Workbook()
    cart_data_sheet = wb.active
    #
    for r in sheets_name.rows:
        cell = sheets_name.cell(row=x, column=1)
        #
        print (cell.value+"         "+str(x))
        #
        x += 1
        #
        card_number = {
            'card_number': cell.value,
        }
        response = requests.get(url=url,params=card_number,headers=headers)
        data_obj = response.json()
        #
        while data_obj['data']['meta']['pagination']['total_pages']:
            card_data = {
                        'card_number': cell.value,
                        'page': data_obj['data']['meta']['pagination']['total_pages']
                    }
            response2 = requests.get(url=url, params=card_data, headers=headers)
            data_obj2 = response2.json()
            #
            if data_obj2['data']['meta']['pagination']['count']==0:
                break
            else:
                #
                print (data_obj['data']['meta']['pagination']['total_pages'])
                #
                data_obj['data']['meta']['pagination']['total_pages'] += 1
                #
                for i2 in range(len(data_obj2['data']['data'])):
                    card_data_list = []
                    card_data_list.append(data_obj2['data']['data'][i2]['card_number'])
                    card_data_list.append(data_obj2['data']['data'][i2]['authorization_date'])
                    card_data_list.append(data_obj2['data']['data'][i2]['merchant_name'])
                    card_data_list.append(data_obj2['data']['data'][i2]['billing_amount'])
                    card_data_list.append(data_obj2['data']['data'][i2]['billing_currency'])
                    card_data_list.append(data_obj2['data']['data'][i2]['status'])
                    card_data_list.append(data_obj2['data']['data'][i2]['fail_reason'])
                    cart_data_sheet.append(card_data_list)
    wb.save('G:\cart_data.xls')
    print ("over+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++over")










