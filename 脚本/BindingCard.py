# -*- coding: utf-8 -*-
"""
Created on Mon Sep  6 16:26:44 2021

@author: Akai
"""
from time import sleep
import pickle
import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import sys
reload(sys)
sys.setdefaultencoding( "utf-8" )

class Card():
    # 初始化
    def __init__(self):
        self.url = "https://business.facebook.com/settings/ad-accounts/262329939096658?business_id=149674206145475"
        xl = openpyxl.load_workbook(r'G:\id.xlsx')
        self.sheets_name = xl.get_sheet_by_name('Sheet2')
        # a=[1,2,3]
        # self.save_variable1(a, r'G:\re.txt')
        a=self.load_variavle(r'G:\re.txt')
        prin
        # 表格游标
        self.x = 0
        # 账号密码游标
        self.y = 0
        # 根据次数切换账号游标
        self.z = 0
        # 账号密码数组
        self.ad_list = ['amarmonar.manush.18', 'amarmonar.manush.19']
        self.pws_list = ['305200', '875837']
        # 储存异常数组
        self.err_list = []
        self.stop_list = []
        # self.binCard()

    # 保存变量方法
    def save_variable1(self, v, filename):
        f = open(filename, 'wb')
        pickle.dump(v, f)
        f.close()
        return filename

    # 读取变量方法
    def load_variavle(self, filename):
        f = open(filename, 'rb')
        r = pickle.load(f)
        f.close()
        return r

    # 登录
    def login(self, ad, pas):
        self.bro = webdriver.Chrome(executable_path=r'C:\ProgramData\Anaconda3\Scripts\chromedriver.exe')
        self.bro.get(self.url)
        self.bro.find_element_by_xpath('//*[@id="email"]').send_keys(ad)
        self.bro.find_element_by_xpath('//*[@id="pass"]').send_keys(pas)
        self.bro.find_element_by_xpath('//*[@id="loginbutton"]').click()
        # self.bro.find_element_by_xpath('//*[@id="approvals_code"]').send_keys(cod)
        # self.bro.find_element_by_xpath('//*[@id="checkpointSubmitButton"]').click()
        # self.bro.find_element_by_xpath('// *[ @ id = "checkpointSubmitButton"]').click()
        self.bro.implicitly_wait(50)

    # 绑卡
    def binCard(self):
        self.login(self.ad_list[self.z], self.pws_list[self.z])
        while 1:
            print (self.y)
            if self.y >= 2:
                self.y = 0
                self.bro.quit()
                self.z += 1
                if self.z >= len(self.ad_list):
                    self.z = 0
                self.login(self.ad_list[self.z], self.pws_list[self.z])

            self.y += 1
            self.x += 1
            cell_id_1 = self.sheets_name.cell(row=self.x, column=1)
            if cell_id_1.value != None:
                # id
                cell_id = self.sheets_name.cell(row=self.x, column=1)
                # 卡号
                cell_name = self.sheets_name.cell(row=self.x, column=2)
                # 安全码
                cell_cod = self.sheets_name.cell(row=self.x, column=3)
                #
                try:
                    loc1 = self.bro.find_element_by_xpath('//*[@placeholder="输入编号/资产名称/商务管理平台名称"]')
                    loc1.send_keys(Keys.CONTROL + 'a')
                    loc1.send_keys(Keys.DELETE)
                    loc1.send_keys(cell_id.value)
                    #
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath('//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]').click()
                    #
                    self.bro.implicitly_wait(30)
                    #
                    Dea = self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]/div[2]/div[3]')
                    dea = Dea.text
                    if dea == '停用的广告帐户':
                        print (cell_id.value + "停用" + cell_name.value)
                        self.stop_list.append(cell_id.value)
                    else:
                        print (cell_id.value + "可用" + cell_name.value)

                        # 下拉点击添加支付按钮
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath('//*[@class="_271o img sp_FqQ3wllOciV sx_669bb6"]').click()
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath('//*[@class="_6ff7"]').click()

                        # 跳转操作页权限
                        self.bro.implicitly_wait(30)
                        self.bro.switch_to_window(self.bro.window_handles[-1])
                        # 点击添加支付按钮
                        self.bro.find_element_by_xpath('//*[@aria-label="添加支付方式"]').click()
                        self.bro.implicitly_wait(30)

                        # 输入账单地址所在国家或地区
                        # 美国下拉按钮
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[1]/div/div/div/div/div/div/label/div/div[2]/div/i').click()
                        self.bro.implicitly_wait(30)
                        # 美国滑动滚动条
                        js = 'var a1=document.getElementsByClassName("j83agx80 cbu4d94t buofh1pr l9j0dhe7")[0].scrollTop=4500'
                        self.bro.execute_script(js)
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[139]').click()
                        # 美元下拉按钮
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[2]/div/div/div/div/div/div/label/div/div[2]/div/i').click()
                        self.bro.implicitly_wait(30)
                        # 美元滑动滚动条
                        js = 'var a2=document.getElementsByClassName("j83agx80 cbu4d94t buofh1pr l9j0dhe7")[0].scrollTop=700'
                        self.bro.execute_script(js)
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[27]').click()
                        # 点击保存按钮
                        sleep(2)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[4]/div/div/div/div[1]/div[1]').click()
                        self.bro.implicitly_wait(30)
                        # 点击继续按钮
                        sleep(2)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[4]/div/div/div/div[1]/div[1]/div/div[1]/div/span/span').click()
                        self.bro.implicitly_wait(30)

                        # 输入卡信息
                        # 持卡人姓名
                        loc2 = self.bro.find_element_by_xpath('//*[@aria-label="持卡人姓名"]')
                        loc2.send_keys("1")
                        # 卡号
                        loc3 = self.bro.find_element_by_xpath('//*[@aria-label="卡号"]')
                        loc3.send_keys(cell_name.value)
                        # 日期
                        loc4 = self.bro.find_element_by_xpath('//*[@aria-label="MM/YY"]')
                        loc4.send_keys("0923")
                        # 安全码
                        loc5 = self.bro.find_element_by_xpath('//*[@aria-label="安全码"]')
                        loc5.send_keys(cell_cod.value)
                        # 保存按钮
                        sleep(2)
                        self.bro.find_element_by_xpath(
                            '//*[@class="d2edcug0 hpfvmrgz qv66sw1b c1et5uql lr9zc1uh a8c37x1j keod5gw0 nxhoafnm aigsh9s9 d3f4x2em fe6kdd0r mau55g9w c8b282yb iv3no6db jq4qci2q a3bd9o3v lrazzd5p bwm1u5wc"]').click()
                        self.bro.implicitly_wait(50)

                        # 输入公司信息
                        # 公司名称
                        loc6 = self.bro.find_element_by_xpath('//*[@aria-label="请添加名称"]')
                        loc6.send_keys('1')
                        # 街道地址第 1 行
                        loc7 = self.bro.find_element_by_xpath('//*[@aria-label="街道地址第 1 行"]')
                        loc7.send_keys('1')
                        # 街道地址第 2 行
                        loc8 = self.bro.find_element_by_xpath('//*[@aria-label="街道地址第 2 行"]')
                        loc8.send_keys('1')
                        # 市/镇
                        loc9 = self.bro.find_element_by_xpath('//*[@aria-label="市/镇"]')
                        loc9.send_keys('1')
                        # 邮编
                        loc10 = self.bro.find_element_by_xpath('//*[@aria-label="邮编"]')
                        loc10.send_keys('00000')
                        # 所在州省
                        self.bro.find_element_by_xpath(
                            '//*[@class="hu5pjgll lzf7d6o1 sp_PYm1K0wZ5y- sx_9c7ed6"]').click()
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[1]').click()
                        # 保存
                        self.bro.find_element_by_xpath('//*[@aria-label="保存"]').click()

                        # 关闭当前标签页
                        self.bro.close()
                        self.bro.switch_to_window(self.bro.window_handles[-1])


                except:
                    print (cell_id.value + '网络错误' + cell_name.value)
                    self.err_list.append(cell_id.value)
                    # self.bro.close()
                    # self.bro.switch_to_window(self.bro.window_handles[-1])
            else:
                print ("停用账户:")
                print (self.stop_list)
                print ("网络异常账户:")
                print (self.err_list)
                self.bro.quit()
                break


if __name__ == "__main__":
    card = Card()

