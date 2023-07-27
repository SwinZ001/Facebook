import global_variable as glv
import os
from time import sleep

import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from tkinter import *

# pyinstaller -F -w UI.py
# 绑卡改名类
class Card():
    def __init__(self,ex,driver):
        print(glv.get("aa"))
        self.rbman_var=glv.get("rbman_var")
        self.text=glv.get("scrolledtext")
        self.start_btn = glv.get("start_btn")
        self.log = ""

        # id表格游标
        self.x = 0

        # 账号游标
        self.use_i = 0

        # 记录异常信息数组
        self.err_list = []
        self.stop_list = []

        # 账号数组
        self.ad_list = []
        # 密码数组
        self.pws_list = []

        # 表格路径值
        self.ex=ex
        # 加载excel文件
        self.load_excel()
        # # 加载驱动文件
        # if getattr(sys,'frozen',False):
        #     chromedriver_path=os.path.join(sys._MEIPASS,"chromedriver.exe")
        #     self.bro=webdriver.Chrome(chromedriver_path)
        # else:
        #     self.bro = webdriver.Chrome(executable_path='F:\python\python3.9.10\Scripts\chromedriver.exe')
        self.bro = webdriver.Chrome(executable_path=driver)
        self.bro.set_window_size(600, 740)
        self.bro.set_window_position(752, 50)

    # 加载excel文件
    def load_excel(self):
        # 读取excel文件
        self.xl = openpyxl.load_workbook(self.ex)
        self.sheets_name = self.xl.get_sheet_by_name('Sheet')
        while 1:
            self.use_i += 1
            cell_use = self.sheets_name.cell(row=self.use_i, column=3)
            if cell_use.value != None:
                self.ad_list.append(self.sheets_name.cell(row=self.use_i, column=3).value)
                self.pws_list.append(self.sheets_name.cell(row=self.use_i, column=4).value)
            else:
                print(str(self.ad_list))
                break

        # 读取Excel中账号使用次数切换游标
        self.y = self.sheets_name.cell(row=1, column=1).value
        # 根据Excel账号次数切换账号游标
        self.z = self.sheets_name.cell(row=2, column=1).value
        self.log = "开始执行..."
        self.text.insert(END, self.log + '\n')
        self.text.update()


    # 登录
    def login(self, url, ad, pas):
        self.bro.get(url)
        self.bro.find_element_by_xpath('//*[@id="email"]').send_keys(ad)
        self.bro.find_element_by_xpath('//*[@id="pass"]').send_keys(pas)
        self.bro.find_element_by_xpath('//*[@id="loginbutton"]').click()
        self.bro.implicitly_wait(30)

    # 操作
    def start_task(self, url):
        # 根据单选按钮的值来判断改名还是绑卡
        if self.rbman_var.get() == 0:
            # 绑卡
            self.login(url,self.ad_list[self.z], self.pws_list[self.z])
            self.bind_card()
        else:
            # 改名
            self.login(url,self.ad_list[0], self.pws_list[0])
            self.rename()

    # 绑卡操作
    def bind_card(self):
        self.log = "账号" + self.ad_list[self.z] + "\n" + "第" + str(self.y) + "次"
        self.text.insert(END, self.log + '\n')
        self.text.update()
        while 1:
            # 账号使用次数大于50（从0开始，实际上是49次），换账号绑卡
            if self.y >= 50:
                self.y = 0
                self.bro.quit()
                self.z += 1
                self.sheets_name["A1"] = self.z
                self.xl.save(self.ex)
                # 账号到最后一个时，切换回第一个
                if self.z >= len(self.ad_list):
                    self.z = 0
                    self.sheets_name["A1"] = self.z
                    self.xl.save(self.ex)
                self.login(self.ad_list[self.z], self.pws_list[self.z])

            self.x += 1
            cell_id_1 = self.sheets_name.cell(row=self.x, column=6)
            if cell_id_1.value != None:
                # 记录账号使用次数
                self.y += 1
                print("账号" + self.ad_list[self.z] +"\n" + "第" + str(self.y) + "次")
                # 更新使用次数
                self.sheets_name["A2"] = self.y
                self.xl.save(self.ex)

                # id
                cell_id = self.sheets_name.cell(row=self.x, column=6)
                # 卡号
                cell_name = self.sheets_name.cell(row=self.x, column=7)
                # 安全码
                cell_cod = self.sheets_name.cell(row=self.x, column=8)
                #
                try:
                    loc1 = self.bro.find_element_by_xpath('//*[@placeholder="ID Masuk / Nama Aset / Nama Bisnis"]')
                    loc1.send_keys(Keys.CONTROL + 'a')
                    loc1.send_keys(Keys.DELETE)
                    loc1.send_keys(cell_id.value)
                    #
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath('//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]').click()
                    #
                    self.bro.implicitly_wait(30)
                    #
                    Dea = self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]/div[2]/div[3]')
                    dea = Dea.text
                    if dea == "":
                        self.log = cell_id.value + "可用" + cell_name.value
                        self.text.insert(END, self.log + '\n')
                        self.text.update()
                        print(self.log)
                        # 下拉点击添加支付按钮
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@id="biz_settings_content_view"]/div/div[4]/div[2]/div/div[3]/div/div[2]/div/div/div/div/div[2]/div/div[1]/div/div[2]/div[2]/div/button/div/i').click()
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath('//*[@class="_6ff7"]').click()

                        # 跳转操作页权限
                        self.bro.implicitly_wait(30)
                        # python2.7以下用switch_to_window，3.0以上用switch_to.window
                        # self.bro.switch_to_window(self.bro.window_handles[-1])
                        self.bro.switch_to.window(self.bro.window_handles[-1])
                        # 点击添加支付按钮
                        self.bro.find_element_by_xpath('//*[@aria-label="Tambahkan Metode Pembayaran"]').click()
                        self.bro.implicitly_wait(30)

                        # 输入账单地址所在国家或地区
                        # 美国下拉按钮
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[1]/div').click()
                        self.bro.implicitly_wait(30)
                        # 美国滑动滚动条
                        for i in range(1, 73):
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[' + str(
                                    i) + ']').send_keys(Keys.DOWN)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[73]').click()
                        self.bro.implicitly_wait(30)
                        # 美元下拉按钮
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[2]/div/div/div/div/div/div/label/div/div[2]/div/i').click()
                        self.bro.implicitly_wait(30)
                        # 美元滑动滚动条
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[7]').click()
                        # 点击保存按钮
                        sleep(1)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[4]/div/div/div/div[1]/div[1]').click()
                        self.bro.implicitly_wait(30)
                        # 点击继续按钮
                        sleep(5)
                        self.bro.find_element_by_xpath('//*[@aria-label="Selanjutnya"]').click()
                        self.bro.implicitly_wait(30)

                        # 输入卡信息
                        # 持卡人姓名
                        loc2 = self.bro.find_element_by_xpath('//*[@aria-label="Nama di kartu"]')
                        loc2.send_keys("1")
                        # 卡号
                        loc3 = self.bro.find_element_by_xpath('//*[@aria-label="Nomor Kartu"]')
                        loc3.send_keys(cell_name.value)
                        # 日期
                        loc4 = self.bro.find_element_by_xpath('//*[@aria-label="BB/TT"]')
                        loc4.send_keys("0224")
                        # 安全码
                        loc5 = self.bro.find_element_by_xpath('//*[@aria-label="CVV"]')
                        loc5.send_keys(cell_cod.value)
                        # 保存按钮
                        sleep(1)
                        self.bro.find_element_by_xpath('//*[@aria-label="Simpan"]').click()
                        self.bro.implicitly_wait(50)

                        # 保存
                        self.bro.find_element_by_xpath('//*[@aria-label="Selesai"]').click()

                        # 关闭当前标签页
                        self.bro.close()
                        self.bro.switch_to.window(self.bro.window_handles[-1])
                    else:
                        self.log = cell_id.value + "停用" + cell_name.value
                        self.text.insert(END, self.log + '\n')
                        self.text.update()
                        print(self.log)
                        self.stop_list.append(cell_id.value)


                except:
                    self.log = cell_id.value + "网络错误" + cell_name.value
                    print(self.log)
                    self.text.insert(END, self.log + '\n')
                    self.text.update()
                    self.err_list.append(cell_id.value)
                    self.bro.close()
                    self.bro.switch_to.window(self.bro.window_handles[-1])
            else:
                self.log = "停用账户:" + str(self.stop_list) + '\n' + "网络异常账户:" + str(self.err_list)+ '\n'
                self.text.insert(END, self.log)
                self.text.update()
                print("停用账户:" + str(self.stop_list) + '\n' + "网络异常账户:" + str(self.err_list))
                # 结束后恢复开始按钮可点击状态
                self.start_btn['state'] = NORMAL
                break

    # 改名操作
    def rename(self):
        while 1:
            self.x += 1
            cell_id_1 = self.sheets_name.cell(row=self.x, column=6)
            if cell_id_1.value != None:
                cell_id = self.sheets_name.cell(row=self.x, column=6)
                cell_name = self.sheets_name.cell(row=self.x, column=9)
                #
                try:
                    loc1 = self.bro.find_element_by_xpath('//*[@placeholder="ID Masuk / Nama Aset / Nama Bisnis"]')
                    loc1.send_keys(Keys.CONTROL + 'a')
                    loc1.send_keys(Keys.DELETE)
                    loc1.send_keys(cell_id.value)
                    #
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]').click()
                    #
                    self.bro.implicitly_wait(30)
                    #
                    Dea = self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]/div[2]/div[3]')
                    dea = Dea.text
                    if dea == "":
                        self.log = cell_id.value + "可用" + cell_name.value
                        self.text.insert(END, self.log + '\n')
                        self.text.update()
                        print(self.log)
                        self.bro.find_element_by_xpath(
                            '//*[@id="biz_settings_content_view"]/div/div[4]/div[2]/div/div[3]/div/div[2]/div/div/div/div/div[2]/div/div[1]/div/div[2]/div[1]/button/div/i').click()
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@class="_2pi2 _6ff6"]/li[3]').click()
                        self.bro.implicitly_wait(30)
                        #
                        loc2 = self.bro.find_element_by_xpath('//*[@placeholder="Masukkan nama akun iklan ini."]')
                        loc2.send_keys(Keys.CONTROL + 'a')
                        loc2.send_keys(Keys.DELETE)
                        loc2.send_keys(cell_name.value)
                        #
                        self.bro.implicitly_wait(30)
                        button = self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div[2]/button')
                        button_type = button.get_attribute('aria-disabled')
                        if button_type == 'false':
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div[2]/button').click()
                            self.bro.implicitly_wait(30)
                            sleep(4)
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div/button').click()
                        else:
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div[1]/button').click()
                            self.log = '已改名'
                            self.text.insert(END, self.log + '\n')
                            self.text.update()
                            print(self.log)
                    else:
                        self.log = cell_id.value + "停用" + cell_name.value
                        self.text.insert(END, self.log + '\n')
                        self.text.update()
                        print(self.log)
                        self.stop_list.append(cell_id.value)

                except:
                    self.log = cell_id.value + "网络错误" + cell_name.value
                    print(self.log)
                    self.err_list.append(cell_id.value)
                    self.text.insert(END, self.log + '\n')
                    self.text.update()
                    print(self.log)
                    # 异常后恢复开始按钮可点击状态
                    # start_btn['state'] = NORMAL
            else:
                self.log = "停用账户:" + str(self.stop_list) + '\n' + "网络异常账户:" + str(self.err_list)+ '\n'
                self.text.insert(END, self.log)
                self.text.update()
                print("停用账户:" + str(self.stop_list) + '\n' + "网络异常账户:" + str(self.err_list))
                # 结束后恢复开始按钮可点击状态
                self.start_btn['state'] = NORMAL
                break


