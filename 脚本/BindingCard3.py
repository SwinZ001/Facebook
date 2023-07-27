# -*- coding: utf-8 -*-
"""
Created on Mon Sep  6 16:26:44 2021

@author: Akai
"""
from time import sleep
import pickle
import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
# import sys
# reload(sys)
# sys.setdefaultencoding('utf8')

class Card():
    # 初始化
    def __init__(self):
        self.url = "https://business.facebook.com/settings/ad-accounts/925247497960475?business_id=2772277696222403"
        xl = openpyxl.load_workbook(r'G:\id.xlsx')
        self.sheets_name = xl.get_sheet_by_name('Sheet2')
        
        #记录账号使用次数，根据次数切换账号游标
        self.cursor_list=self.load_variavle(r'G:\re.txt')
        
        # 表格游标
        self.x=0
        # 账号使用次数切换游标
        self.y=self.cursor_list[0]
        print("次数"+str(self.y))
        # 根据次数切换账号游标
        self.z=self.cursor_list[1]
        print("账号"+str(self.z))
        # 账号数组
        self.ad_list = ['shivkumar.kasdekar.7967',
                        'malik.azmatshahzad',
                        'mk.mahabubkhan.98',
                        'loreto.legaspi.94',
                        'AssAsssQ',
                        'eduardo.echeverria.71271',
                        'onder.dominic.9',
                        'onder.dominic.10',
                        'gaga.afghan',
                        'ameer.haider.9028'
                        ]
        # 密码数组
        self.pws_list = ['dscvssa -89',
                         'dscvssa -90',
                         'dscvssa -91',
                         'swinz123456',
                         '1011ysy.520',
                         '1011ysy.520',
                         'mukasa1',
                         '573977881',
                         '1011ysy.520',
                         'as_dsdad..69008'
                          ]
        # 储存异常数组
        self.err_list = []
        self.stop_list = []
        self.binCard()

        
        
    #保存变量方法 
    def save_variable1(self,v,filename):
        f=open(filename,'wb')
        pickle.dump(v,f,0)
        f.close()
        return filename
    #读取变量方法
    def load_variavle(self,filename):
        f=open(filename,'rb')
        r=pickle.load(f)
        f.close()
        return r
  

    # 登录
    def login(self,ad,pas):
        self.bro = webdriver.Chrome(executable_path='F:\python\python3.9.10\Scripts\chromedriver.exe')
        self.bro.get(self.url)
        self.bro.find_element_by_xpath('//*[@id="email"]').send_keys(ad)
        self.bro.find_element_by_xpath('//*[@id="pass"]').send_keys(pas)
        self.bro.find_element_by_xpath('//*[@id="loginbutton"]').click()
        # self.bro.find_element_by_xpath('//*[@id="approvals_code"]').send_keys(cod)
        # self.bro.find_element_by_xpath('//*[@id="checkpointSubmitButton"]').click()
        # self.bro.find_element_by_xpath('// *[ @ id = "checkpointSubmitButton"]').click()
        self.bro.implicitly_wait(50)


    # 绑卡
    def binCard(self):
        self.login(self.ad_list[self.z], self.pws_list[self.z])
        while 1:
            # 账号使用次数大于50（从0开始，实际上是49次），换账号绑卡
            if self.y >=50:
                self.y=0
                self.bro.quit()
                self.z += 1
                self.cursor_list[1]=self.z
                # 账号到最后一个时，切换回第一个
                if self.z >= len(self.ad_list):
                    self.z = 0
                    self.cursor_list[1]=self.z
                self.login(self.ad_list[self.z], self.pws_list[self.z])
    
            self.x+=1
            cell_id_1 = self.sheets_name.cell(row=self.x,column=1)
            if cell_id_1.value != None:
                # 记录账号使用次数
                self.y+=1
                print ("账号"+self.ad_list[self.z]+"第"+str(self.y)+"次")
                # 更新使用次数
                self.cursor_list[0]=self.y
                self.save_variable1(self.cursor_list,r'G:\re.txt')
                
                # id
                cell_id = self.sheets_name.cell(row=self.x, column=1)
                # 卡号
                cell_name = self.sheets_name.cell(row=self.x, column=2)
                # 安全码
                cell_cod = self.sheets_name.cell(row=self.x, column=3)
               #
                # try:
                loc1 = self.bro.find_element_by_xpath('//*[@placeholder="ID Masuk / Nama Aset / Nama Bisnis"]')
                loc1.send_keys(Keys.CONTROL + 'a')
                loc1.send_keys(Keys.DELETE)
                loc1.send_keys(cell_id.value)
                #
                self.bro.implicitly_wait(30)
                self.bro.find_element_by_xpath('//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]').click()
                #
                self.bro.implicitly_wait(30)
                #
                Dea = self.bro.find_element_by_xpath('//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]/div[2]/div[3]')
                dea = Dea.text
                if dea == 'Akun Iklan yang Dinonaktifkan':
                    print ("\033[31m" + cell_id.value + "停用" + cell_name.value + "\033[0m")
                    self.stop_list.append(cell_id.value)
                else:
                    print (cell_id.value + "可用" + cell_name.value)
                    # 下拉点击添加支付按钮
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath('//*[@id="biz_settings_content_view"]/div/div[4]/div[2]/div/div[3]/div/div[2]/div/div/div/div/div[2]/div/div[1]/div/div[2]/div[2]/div/button/div/i').click()
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath('//*[@class="_6ff7"]').click()

                    # 跳转操作页权限
                    self.bro.implicitly_wait(30)
                    self.bro.switch_to_window(self.bro.window_handles[-1])
                    # 点击添加支付按钮
                    self.bro.find_element_by_xpath('//*[@aria-label="Tambahkan Metode Pembayaran"]').click()
                    self.bro.implicitly_wait(30)





                    # 输入账单地址所在国家或地区
                    # 美国下拉按钮
                    self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[1]/div').click()
                    self.bro.implicitly_wait(30)
                    # 美国滑动滚动条
                    for i in range(1, 73):
                        self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[' + str(i) + ']').send_keys(Keys.DOWN)
                    self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[73]').click()
                    self.bro.implicitly_wait(30)
                    # # 美元滑动滚动条
                    # for i2 in range(1, 27):
                    #     self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[' + str(i2) + ']').send_keys(Keys.DOWN)
                    # self.bro.implicitly_wait(30)
                    # self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[27]').click()

                    # # 输入账单地址所在国家或地区
                    # # 美国下拉按钮
                    # self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[1]/div').click()
                    # self.bro.implicitly_wait(30)
                    # # 美国滑动滚动条
                    # self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[5]').click()
                    # self.bro.implicitly_wait(30)
                    # 美元下拉按钮
                    self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[2]/div/div/div/div/div/div/label/div/div[2]/div/i').click()
                    self.bro.implicitly_wait(30)
                    # 美元滑动滚动条
                    self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[7]').click()
                    # 点击保存按钮
                    sleep(1)
                    self.bro.find_element_by_xpath('//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[4]/div/div/div/div[1]/div[1]').click()
                    self.bro.implicitly_wait(30)
                    # 点击继续按钮
                    sleep(5)
                    self.bro.find_element_by_xpath('//*[@aria-label="Selanjutnya"]').click()
                    self.bro.implicitly_wait(30)


                   # 输入卡信息
                    # 持卡人姓名
                    loc2=self.bro.find_element_by_xpath('//*[@aria-label="Nama di kartu"]')
                    loc2.send_keys("1")
                    # 卡号
                    loc3=self.bro.find_element_by_xpath('//*[@aria-label="Nomor Kartu"]')
                    loc3.send_keys(cell_name.value)
                    # 日期
                    loc4=self.bro.find_element_by_xpath('//*[@aria-label="BB/TT"]')
                    loc4.send_keys("0224")
                    # 安全码
                    loc5=self.bro.find_element_by_xpath('//*[@aria-label="CVV"]')
                    loc5.send_keys(cell_cod.value)
                    # 保存按钮
                    sleep(1)
                    self.bro.find_element_by_xpath('//*[@aria-label="Simpan"]').click()
                    self.bro.implicitly_wait(50)

                    # 保存
                    self.bro.find_element_by_xpath('//*[@aria-label="Selesai"]').click()


                   # # 输入公司信息
                   #  # 公司名称
                   #  loc6 = self.bro.find_element_by_xpath('//*[@aria-label="Tambahkan nama"]')
                   #  loc6.send_keys('1')
                   #  # 街道地址第 1 行
                   #  loc7 = self.bro.find_element_by_xpath('//*[@aria-label="Alamat 1"]')
                   #  loc7.send_keys('1')
                   #  # 街道地址第 2 行
                   #  loc8 = self.bro.find_element_by_xpath('//*[@aria-label="Alamat 2"]')
                   #  loc8.send_keys('1')
                   #  # 市/镇
                   #  loc9 = self.bro.find_element_by_xpath('//*[@aria-label="Kota"]')
                   #  loc9.send_keys('1')
                   #  # 邮编
                   #  loc10 = self.bro.find_element_by_xpath('//*[@aria-label="Kode Pos"]')
                   #  loc10.send_keys('00000')
                   #  # 所在州省
                   #  self.bro.find_element_by_xpath(
                   #      '//*[@class="hu5pjgll lzf7d6o1 sp_PYm1K0wZ5y- sx_9c7ed6"]').click()
                   #  self.bro.implicitly_wait(30)
                   #  self.bro.find_element_by_xpath(
                   #      '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[1]').click()
                    # # 保存
                    # self.bro.find_element_by_xpath('//*[@aria-label="Simpan"]').click()






                    # 关闭当前标签页
                    self.bro.close()
                    self.bro.switch_to_window(self.bro.window_handles[-1])
                        
                # except:
                #     print ("\033[32m" + cell_id.value + "网络错误" + cell_name.value + "\033[0m")
                #     self.err_list.append(cell_id.value)
                #     self.bro.close()
                #     self.bro.switch_to_window(self.bro.window_handles[-1])
            else:
                print ("停用账户:")
                print (self.stop_list)
                print ("网络异常账户:")
                print (self.err_list)
                # self.bro.quit()
                break
        
if __name__ == "__main__":
    card=Card()
    
    