from selenium.common.exceptions import NoSuchElementException

import global_variable as glv
# import os
from time import sleep
from sqlTest import Db_utils
import openpyxl as openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from tkinter import *

# pyinstaller -F -w UI.py
# 绑卡改名类
class Card():
    def __init__(self,ex,driver):
        # 连接数据库
        self.db = Db_utils(host="localhost", database="test", user="root", password="123456")
        # 获取全局变量
        self.rbman_var=glv.get("rbman_var")
        self.text=glv.get("scrolledtext")
        self.start_btn = glv.get("start_btn")
        self.log = ""

        # id表格游标
        self.x = 0

        # 账号游标
        self.use_i = 0

        # 记录异常信息数组
        self.err_list = []
        self.stop_list = []

        # 账号数组
        self.ad_list = []
        # 密码数组
        self.pws_list = []

        # 表格路径值
        self.ex=ex
        # 加载excel文件
        self.load_excel()
        # # 加载驱动文件
        # if getattr(sys,'frozen',False):
        #     chromedriver_path=os.path.join(sys._MEIPASS,"chromedriver.exe")
        #     self.bro=webdriver.Chrome(chromedriver_path)
        # else:
        #     self.bro = webdriver.Chrome(executable_path='F:\python\python3.9.10\Scripts\chromedriver.exe')
        self.bro = webdriver.Chrome(executable_path=driver)
        self.bro.set_window_size(600, 740)
        self.bro.set_window_position(752, 50)

    # 加载excel文件
    def load_excel(self):
        # 读取excel文件
        self.xl = openpyxl.load_workbook(self.ex)
        self.sheets_name = self.xl.get_sheet_by_name('Sheet')
        while 1:
            self.use_i += 1
            cell_use = self.sheets_name.cell(row=self.use_i, column=3)
            if cell_use.value != None:
                self.ad_list.append(self.sheets_name.cell(row=self.use_i, column=3).value)
                self.pws_list.append(self.sheets_name.cell(row=self.use_i, column=4).value)
            else:
                print(str(self.ad_list))
                break

        # 读取Excel中账号使用次数切换游标
        self.y = self.sheets_name.cell(row=1, column=1).value
        # 根据Excel账号次数切换账号游标
        self.z = self.sheets_name.cell(row=2, column=1).value
        self.set_log("开始执行.."+"\n")


    # 登录
    def login(self, url, ad, pas):
        self.bro.get(url)
        self.bro.find_element_by_xpath('//*[@id="email"]').send_keys(ad)
        self.bro.find_element_by_xpath('//*[@id="pass"]').send_keys(pas)
        self.bro.find_element_by_xpath('//*[@id="loginbutton"]').click()
        self.bro.implicitly_wait(30)
    # 输出log
    def set_log(self,str):
        self.log = str
        self.text.insert(END, self.log + '\n')
        self.text.see(END)
        self.text.update()
        print(self.log)
    # 操作
    def start_task(self, url):
        # 根据单选按钮的值来判断改名还是绑卡
        if self.rbman_var.get() == 0:
            # 绑卡
            self.login(url,self.ad_list[self.z], self.pws_list[self.z])
            self.bind_card()
        elif self.rbman_var.get() == 1:
            # 改名
            self.login(url,self.ad_list[0], self.pws_list[0])
            self.rename()
        elif self.rbman_var.get() == 2:
            self.login(url, self.ad_list[0], self.pws_list[0])
            self.get_userData()
    # 绑卡操作
    def bind_card(self):
        self.set_log("开始执行绑卡..")
        while 1:
            # 账号使用次数大于50（从0开始，实际上是49次），换账号绑卡
            if self.y >= 50:
                self.y = 0
                self.bro.quit()
                self.z += 1
                self.sheets_name["A1"] = self.z
                self.xl.save(self.ex)
                # 账号到最后一个时，切换回第一个
                if self.z >= len(self.ad_list):
                    self.z = 0
                    self.sheets_name["A1"] = self.z
                    self.xl.save(self.ex)
                self.login(self.ad_list[self.z], self.pws_list[self.z])

            self.x += 1
            cell_id_1 = self.sheets_name.cell(row=self.x, column=6)
            if cell_id_1.value != None:
                # 记录账号使用次数
                self.y += 1
                self.set_log("账号" + self.ad_list[self.z] + "\n" + "第" + str(self.y) + "次")
                # 更新使用次数
                self.sheets_name["A2"] = self.y
                self.xl.save(self.ex)

                # id
                cell_id = self.sheets_name.cell(row=self.x, column=6)
                # 卡号
                cell_name = self.sheets_name.cell(row=self.x, column=7)
                # 安全码
                cell_cod = self.sheets_name.cell(row=self.x, column=8)
                #
                try:
                    loc1 = self.bro.find_element_by_xpath('//*[@placeholder="ID Masuk / Nama Aset / Nama Bisnis"]')
                    loc1.send_keys(Keys.CONTROL + 'a')
                    loc1.send_keys(Keys.DELETE)
                    loc1.send_keys(cell_id.value)
                    #
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath('//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]').click()
                    #
                    self.bro.implicitly_wait(30)
                    #
                    Dea = self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]/div[2]/div[3]')
                    dea = Dea.text
                    if dea == "":
                        self.set_log(cell_id.value + "可用" + cell_name.value)
                        # 下拉点击添加支付按钮
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@id="biz_settings_content_view"]/div/div[4]/div[2]/div/div[3]/div/div[2]/div/div/div/div/div[2]/div/div[1]/div/div[2]/div[2]/div/button/div/i').click()
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath('//*[@class="_6ff7"]').click()

                        # 跳转操作页权限
                        self.bro.implicitly_wait(30)
                        # python2.7以下用switch_to_window，3.0以上用switch_to.window
                        # self.bro.switch_to_window(self.bro.window_handles[-1])
                        self.bro.switch_to.window(self.bro.window_handles[-1])
                        # 点击添加支付按钮
                        self.bro.find_element_by_xpath('//*[@aria-label="Tambahkan Metode Pembayaran"]').click()
                        self.bro.implicitly_wait(30)

                        # 输入账单地址所在国家或地区
                        # 美国下拉按钮
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[1]/div').click()
                        self.bro.implicitly_wait(30)
                        # 美国滑动滚动条
                        for i in range(1, 73):
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[' + str(
                                    i) + ']').send_keys(Keys.DOWN)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[73]').click()
                        self.bro.implicitly_wait(30)
                        # 美元下拉按钮
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[3]/div/div[1]/div[2]/div[2]/div/div/div/div/div/div/label/div/div[2]/div/i').click()
                        self.bro.implicitly_wait(30)
                        # 美元滑动滚动条
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div[1]/div/div[7]').click()
                        # 点击保存按钮
                        sleep(1)
                        self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[7]/div[2]/div[1]/div[1]/div[1]/div/div/div/div/div/div[1]/div[4]/div/div/div/div[1]/div[1]').click()
                        self.bro.implicitly_wait(30)
                        # 点击继续按钮
                        sleep(5)
                        self.bro.find_element_by_xpath('//*[@aria-label="Selanjutnya"]').click()
                        self.bro.implicitly_wait(30)

                        # 输入卡信息
                        # 持卡人姓名
                        loc2 = self.bro.find_element_by_xpath('//*[@aria-label="Nama di kartu"]')
                        loc2.send_keys("1")
                        # 卡号
                        loc3 = self.bro.find_element_by_xpath('//*[@aria-label="Nomor Kartu"]')
                        loc3.send_keys(cell_name.value)
                        # 日期
                        loc4 = self.bro.find_element_by_xpath('//*[@aria-label="BB/TT"]')
                        loc4.send_keys("0224")
                        # 安全码
                        loc5 = self.bro.find_element_by_xpath('//*[@aria-label="CVV"]')
                        loc5.send_keys(cell_cod.value)
                        # 保存按钮
                        sleep(1)
                        self.bro.find_element_by_xpath('//*[@aria-label="Simpan"]').click()
                        self.bro.implicitly_wait(50)

                        # 保存
                        self.bro.find_element_by_xpath('//*[@aria-label="Selesai"]').click()

                        # 关闭当前标签页
                        self.bro.close()
                        self.bro.switch_to.window(self.bro.window_handles[-1])
                    else:
                        self.set_log(cell_id.value + "停用" + cell_name.value)
                        self.stop_list.append(cell_id.value)


                except:
                    self.set_log(cell_id.value + "网络错误" + cell_name.value)
                    self.err_list.append(cell_id.value)
                    self.bro.close()
                    self.bro.switch_to.window(self.bro.window_handles[-1])
            else:
                self.set_log("停用账户:" + str(self.stop_list) + '\n' + "网络异常账户:" + str(self.err_list)+"\n")
                # 结束后恢复开始按钮可点击状态
                # self.start_btn['state'] = NORMAL
                break

    # 改名操作
    def rename(self):
        self.set_log("开始执行改名..")
        while 1:
            self.x += 1
            cell_id_1 = self.sheets_name.cell(row=self.x, column=6)
            if cell_id_1.value != None:
                cell_id = self.sheets_name.cell(row=self.x, column=6)
                cell_name = self.sheets_name.cell(row=self.x, column=9)
                #
                try:
                    loc1 = self.bro.find_element_by_xpath('//*[@placeholder="ID Masuk / Nama Aset / Nama Bisnis"]')
                    loc1.send_keys(Keys.CONTROL + 'a')
                    loc1.send_keys(Keys.DELETE)
                    loc1.send_keys(cell_id.value)
                    #
                    self.bro.implicitly_wait(30)
                    self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]').click()
                    #
                    self.bro.implicitly_wait(30)
                    #
                    Dea = self.bro.find_element_by_xpath(
                        '//*[@class="_25b6 _21op _2pi9 _3qn7 _61-3 _2fyi _3qng"]/div[2]/div[3]')
                    # 获取文本值
                    dea = Dea.text
                    if dea == "":
                        self.set_log(cell_id.value + "可用" + cell_name.value)
                        self.bro.find_element_by_xpath(
                            '//*[@id="biz_settings_content_view"]/div/div[4]/div[2]/div/div[3]/div/div[2]/div/div/div/div/div[2]/div/div[1]/div/div[2]/div[1]/button/div/i').click()
                        self.bro.implicitly_wait(30)
                        self.bro.find_element_by_xpath(
                            '//*[@class="_2pi2 _6ff6"]/li[3]').click()
                        self.bro.implicitly_wait(30)
                        #
                        loc2 = self.bro.find_element_by_xpath('//*[@placeholder="Masukkan nama akun iklan ini."]')
                        loc2.send_keys(Keys.CONTROL + 'a')
                        loc2.send_keys(Keys.DELETE)
                        loc2.send_keys(cell_name.value)
                        #
                        self.bro.implicitly_wait(30)
                        button = self.bro.find_element_by_xpath(
                            '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div[2]/button')
                        # 获取属性值
                        button_type = button.get_attribute('aria-disabled')
                        if button_type == 'false':
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div[2]/button').click()
                            self.bro.implicitly_wait(30)
                            sleep(4)
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div/button').click()
                        else:
                            self.bro.find_element_by_xpath(
                                '//*[@id="facebook"]/body/div[6]/div[2]/div/div/div/div/div/div[2]/div[3]/span[2]/div/div[1]/button').click()
                            self.set_log('已改名')
                    else:
                        self.set_log(cell_id.value + "停用" + cell_name.value)
                        self.stop_list.append(cell_id.value)

                except:
                    self.set_log(cell_id.value + "网络错误" + cell_name.value)
                    self.err_list.append(cell_id.value)
                    # 异常后恢复开始按钮可点击状态
                    # start_btn['state'] = NORMAL
            else:
                self.set_log("停用账户:" + str(self.stop_list) + '\n' + "网络异常账户:" + str(self.err_list)+ '\n')
                # 结束后恢复开始按钮可点击状态
                # self.start_btn['state'] = NORMAL
                break

    # 爬取用户信息
    def get_userData(self):
        # # 第一种方法，连续点击数据组件，点击一个获取一个数据，比较费时
        # # f=open("G:\data.csv",mode="a",encoding='utf_8_sig')
        # self.set_log("开始执行爬取用户数据..")
        # # 切换组件参数
        # click_i=0
        #
        # line_i=0
        # # 遍历列表参数
        # list_i=0
        # us_name_list = []
        # us_type_list = []
        # us_state_list = []
        # self.set_log("数据爬取中..")
        # while 1:
        #     click_i += 1
        #     try:
        #         self.bro.find_element_by_xpath('//*[@class="uiScrollableAreaContent"]/div/div[1]/div[' + str(click_i) + ']').click()
        #         use_data = self.bro.find_element_by_xpath('//*[@class="uiScrollableAreaContent"]/div/div[1]/div[' + str(click_i) + ']').find_elements_by_class_name("ellipsis")
        #         for user in use_data:
        #             line_i += 1
        #             if line_i == 1:
        #                 us_name_list.append(user.text)
        #             elif line_i == 2:
        #                 us_type_list.append(user.text)
        #             elif line_i == 3:
        #                 us_state_list.append(user.text)
        #                 self.set_log("us_name:"+str(us_name_list[list_i]) + "\n"
        #                              +"us_type:"+str(us_type_list[list_i]) + "\n"
        #                              +"us_state:"+str(us_state_list[list_i]) + "\n")
        #                 # 向数据库添加数据
        #                 self.db.sql_add("insert into userdata values(null,%s,%s,%s)",(us_name_list[list_i],us_type_list[list_i],us_state_list[list_i]))
        #                 print(str(us_name_list[list_i]) + "\n" + str(us_type_list[list_i]) + "\n" + str(us_state_list[list_i]) + "\n")
        #                 list_i+=1
        #                 line_i = 0
        #
        #     except NoSuchElementException as e:
        #         self.set_log("数据爬取完成"+"\n")
        #         print("结束"+str(e))
        #         self.db.closs_db()
        #         break





        self.set_log("开始执行爬取用户数据..")
        # 第二种方法(滚动条滑动加载出全部标签后再进行全部标签数据爬取)
        temp_height = 0
        x = 1100
        y = 1100
        # 换行参数
        line_i = 0
        # 遍历列表参数
        list_i=0
        us_name_list = []
        us_type_list = []
        us_state_list = []
        # f=open("G:\data.csv",mode="a",encoding='utf_8_sig')
        self.set_log("下拉加载滚动条..")
        self.bro.implicitly_wait(30)
        while True:
            # js语句，动态改变滑动值进行滑动.format()
            js1 = "var q=document.getElementsByClassName('uiScrollableAreaWrap scrollable')[0].scrollTop={}".format(x)
            self.bro.execute_script(js1)
            sleep(2)
            x += y
            check_height = self.bro.execute_script(
                "return document.getElementsByClassName('uiScrollableAreaWrap scrollable')[0].scrollTop;")
            if check_height == temp_height:
                break
            temp_height = check_height
            self.set_log(str(temp_height))

        sleep(2)
        user_data = self.bro.find_element_by_xpath(
            '//*[@class="uiScrollableAreaWrap scrollable"]').find_elements_by_class_name("ellipsis")
        self.set_log("数据爬取中..")
        for user in user_data:
            line_i += 1
            if line_i == 1:
                us_name_list.append(user.text)
                # f.write(user.text)
                # f.write(",")
            elif line_i == 2:
                us_type_list.append(user.text)
            elif line_i == 3:
                us_state_list.append(user.text)
                self.set_log("us_name:" + str(us_name_list[list_i]) + "\n"
                             + "us_type:" + str(us_type_list[list_i]) + "\n"
                             + "us_state:" + str(us_state_list[list_i]) + "\n")
                # 向数据库添加数据
                self.db.sql_add("insert into data values(null,%s,%s,%s)",
                                (us_name_list[list_i], us_type_list[list_i], us_state_list[list_i]))
                print(str(us_name_list[list_i]) + "\n" + str(us_type_list[list_i]) + "\n" + str(
                    us_state_list[list_i]) + "\n")
                list_i += 1
                line_i = 0
        self.set_log("数据入库成功"+"\n")
        self.db.closs_db()
        #


